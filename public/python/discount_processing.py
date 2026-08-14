from __future__ import annotations

import argparse
import re
import unicodedata
from collections import defaultdict
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


# Tên dùng để nhận diện dòng chiết khấu.
DISCOUNT_ITEM_NAME = "Chiết khấu hàng hóa"

# Tiền VND thường không có số lẻ.
# Đổi thành 2 nếu file có phần thập phân.
MONEY_DECIMALS = 0


# Các cột bắt buộc phải có trong file.
REQUIRED_HEADERS = (
    "Số HĐ",
    "Tên hàng",
    "Số lượng",
    "Đơn giá",
    "Chiết khấu",
    "Doanh thu",
    "Thuế suất",
    "Tiền thuế",
    "Thành tiền",
)
# Các cột dùng để xác định duy nhất một hóa đơn.
OPTIONAL_KEY_HEADERS = (
    "Ký hiệu mẫu",
    "Ký hiệu",
    "Mã tra cứu",
)


@dataclass
class InvoiceResult:
    invoice: str
    discount: Decimal
    allocated: Decimal
    remaining: Decimal
    affected_products: int


def normalize_text(value: Any) -> str:
    """
    Chuẩn hóa chuỗi để so sánh:
    - Bỏ khoảng trắng đầu cuối.
    - Gom nhiều khoảng trắng thành một.
    - Không phân biệt chữ hoa và chữ thường.
    """

    if value is None:
        return ""

    text = unicodedata.normalize("NFC", str(value))
    text = re.sub(r"\s+", " ", text.strip())

    return text.casefold()


def key_text(value: Any) -> str:
    """
    Chuyển giá trị thành chuỗi để dùng làm khóa hóa đơn.
    Ví dụ 3669.0 sẽ được chuyển thành 3669.
    """

    if value is None:
        return ""

    if isinstance(value, float) and value.is_integer():
        return str(int(value))

    if isinstance(value, Decimal) and value == value.to_integral_value():
        return str(int(value))

    return str(value).strip()


def to_decimal(value: Any) -> Decimal | None:
    """
    Chuyển dữ liệu Excel thành Decimal.

    Hỗ trợ:
    - 1260000
    - 1,260,000
    - 1.260.000
    - (1,260,000)
    - 10%
    """

    if value is None or value == "":
        return None

    if isinstance(value, Decimal):
        return value

    if isinstance(value, bool):
        return Decimal(int(value))

    if isinstance(value, int):
        return Decimal(value)

    if isinstance(value, float):
        return Decimal(str(value))

    text = str(value).strip()

    if not text:
        return None

    # Excel có thể biểu diễn số âm dạng (1,260,000).
    negative = text.startswith("(") and text.endswith(")")

    if negative:
        text = text[1:-1]

    text = (
        text.replace("\u00a0", "")
        .replace(" ", "")
        .replace("₫", "")
        .replace("VND", "")
        .replace("vnd", "")
        .replace("%", "")
    )

    # Chỉ giữ số và ký tự phân cách.
    text = re.sub(r"[^0-9,\.\-+]", "", text)

    if not text or text in {"-", "+", ".", ","}:
        return None

    # Có cả dấu phẩy và dấu chấm.
    if "," in text and "." in text:
        # Dấu xuất hiện cuối cùng được xem là dấu thập phân.
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")

    # Chỉ có dấu phẩy.
    elif "," in text:
        if text.count(",") > 1:
            # Ví dụ 1,260,000.
            text = text.replace(",", "")
        else:
            before, after = text.split(",", 1)

            if len(after) == 3 and before.lstrip("+-").isdigit():
                # Ví dụ 160,000.
                text = before + after
            else:
                # Ví dụ 10,5.
                text = before + "." + after

    # Chỉ có dấu chấm.
    elif "." in text:
        if text.count(".") > 1:
            # Ví dụ 1.260.000.
            text = text.replace(".", "")
        else:
            before, after = text.split(".", 1)

            if len(after) == 3 and before.lstrip("+-").isdigit():
                # Ví dụ 160.000.
                text = before + after

    try:
        result = Decimal(text)
    except InvalidOperation:
        return None

    if negative:
        result = -result

    return result

def tax_rate_key(value: Any) -> tuple[str, str]:
    """
    Chuẩn hóa thuế suất để so sánh.

    Ví dụ:
    - 10 và "10%" được xem là giống nhau.
    - 0.1 trong Excel cũng được xem là 10%.
    - Các giá trị chữ như KCT được so sánh theo chuỗi.
    """

    rate = to_decimal(value)

    if rate is not None:
        # Excel có thể lưu 10% thành 0.1.
        if rate != 0 and abs(rate) < 1:
            rate *= Decimal("100")

        rate = rate.quantize(
            Decimal("0.0001")
        ).normalize()

        return (
            "number",
            str(rate),
        )

    return (
        "text",
        normalize_text(value),
    )


def round_money(value: Decimal) -> Decimal:
    """Làm tròn số tiền."""

    if MONEY_DECIMALS == 0:
        quantum = Decimal("1")
    else:
        quantum = Decimal("1").scaleb(-MONEY_DECIMALS)

    return value.quantize(
        quantum,
        rounding=ROUND_HALF_UP,
    )


def excel_number(value: Decimal) -> int | float:
    """Chuyển Decimal thành kiểu dữ liệu Excel có thể ghi."""

    value = round_money(value)

    if value == value.to_integral_value():
        return int(value)

    return float(value)


def read_cell(
    ws_values: Worksheet,
    ws_write: Worksheet,
    row: int,
    col: int,
) -> Any:
    """
    Ưu tiên lấy kết quả đã tính của công thức.

    ws_values:
        Workbook được mở với data_only=True.

    ws_write:
        Workbook giữ nguyên công thức và định dạng.
    """

    calculated_value = ws_values.cell(
        row=row,
        column=col,
    ).value

    if calculated_value is not None:
        return calculated_value

    raw_value = ws_write.cell(
        row=row,
        column=col,
    ).value

    # Nếu là công thức nhưng không có cached value thì không thể tự tính.
    if isinstance(raw_value, str) and raw_value.startswith("="):
        return None

    return raw_value


def find_header_row(
    ws: Worksheet,
    max_search_rows: int = 100,
) -> tuple[int, dict[str, int]] | None:
    """
    Tìm dòng chứa tiêu đề bảng.

    Điều này hữu ích khi phía trên bảng có:
    - Tên báo cáo.
    - Kỳ báo cáo.
    - Tên công ty.
    - Ghi chú.
    """

    normalized_required = {
        normalize_text(header): header
        for header in REQUIRED_HEADERS
    }

    normalized_optional = {
        normalize_text(header): header
        for header in OPTIONAL_KEY_HEADERS
    }

    last_search_row = min(
        ws.max_row,
        max_search_rows,
    )

    for row in range(1, last_search_row + 1):
        normalized_to_col: dict[str, int] = {}

        for col in range(1, ws.max_column + 1):
            normalized = normalize_text(
                ws.cell(
                    row=row,
                    column=col,
                ).value
            )

            if normalized and normalized not in normalized_to_col:
                normalized_to_col[normalized] = col

        has_all_required_headers = all(
            header in normalized_to_col
            for header in normalized_required
        )

        if not has_all_required_headers:
            continue

        columns = {
            canonical: normalized_to_col[normalized]
            for normalized, canonical in normalized_required.items()
        }

        for normalized, canonical in normalized_optional.items():
            if normalized in normalized_to_col:
                columns[canonical] = normalized_to_col[normalized]

        return row, columns

    return None


def tax_ratio(
    rate_value: Any,
    old_revenue: Decimal,
    old_tax: Decimal,
) -> Decimal:
    """
    Chuyển thuế suất thành tỷ lệ.

    Ví dụ:
    - Excel lưu 10% thành 0.1.
    - File dữ liệu có thể lưu thành 10.
    """

    rate = to_decimal(rate_value)

    if rate is not None:
        # Trường hợp Excel lưu 10% thành 0.1.
        if Decimal("0") < abs(rate) < Decimal("1"):
            return rate

        # Trường hợp dữ liệu lưu thuế suất là 10.
        return rate / Decimal("100")

    # Không đọc được cột Thuế suất thì suy ra từ số tiền cũ.
    if old_revenue != 0:
        return old_tax / old_revenue

    return Decimal("0")


def build_invoice_key(
    ws_values: Worksheet,
    ws_write: Worksheet,
    row: int,
    columns: dict[str, int],
) -> tuple[str, ...] | None:
    """
    Tạo khóa nhóm hóa đơn.

    Ưu tiên:
    1. Mã tra cứu.
    2. Ký hiệu mẫu + Ký hiệu + Số HĐ.
    """

    invoice_number = key_text(
        read_cell(
            ws_values,
            ws_write,
            row,
            columns["Số HĐ"],
        )
    )

    if not invoice_number:
        return None

    lookup_code = ""

    if "Mã tra cứu" in columns:
        lookup_code = key_text(
            read_cell(
                ws_values,
                ws_write,
                row,
                columns["Mã tra cứu"],
            )
        )

    # Mã tra cứu thường là khóa duy nhất của hóa đơn.
    if lookup_code:
        return (
            "lookup",
            lookup_code,
        )

    template_symbol = ""
    invoice_symbol = ""

    if "Ký hiệu mẫu" in columns:
        template_symbol = key_text(
            read_cell(
                ws_values,
                ws_write,
                row,
                columns["Ký hiệu mẫu"],
            )
        )

    if "Ký hiệu" in columns:
        invoice_symbol = key_text(
            read_cell(
                ws_values,
                ws_write,
                row,
                columns["Ký hiệu"],
            )
        )

    return (
        "invoice",
        template_symbol,
        invoice_symbol,
        invoice_number,
    )


def invoice_label(key: tuple[str, ...]) -> str:
    """Tạo nội dung dùng để in log."""

    if key[0] == "lookup":
        return f"Mã tra cứu {key[1]}"

    _, template_symbol, invoice_symbol, invoice_number = key

    parts = [
        part
        for part in (
            template_symbol,
            invoice_symbol,
            invoice_number,
        )
        if part
    ]

    return " / ".join(parts)


def set_money(
    ws: Worksheet,
    row: int,
    col: int,
    value: Decimal,
) -> None:
    """Ghi số tiền vào ô Excel."""

    ws.cell(
        row=row,
        column=col,
    ).value = excel_number(value)



def process_invoice(
    ws_values: Worksheet,
    ws_write: Worksheet,
    rows: list[int],
    columns: dict[str, int],
    invoice: str,
) -> InvoiceResult | None:
    """
    Xử lý một hóa đơn.

    Quy tắc:
    - Tìm dòng có Tên hàng = Chiết khấu hàng hóa.
    - Chỉ phân bổ chiết khấu cho mặt hàng cùng thuế suất.
    - Trong từng thuế suất, ưu tiên dòng có:
      Số lượng × Đơn giá lớn nhất.
    """

    normalized_discount_name = normalize_text(
        DISCOUNT_ITEM_NAME
    )

    discount_rows: list[int] = []

    # Tìm tất cả dòng chiết khấu trong hóa đơn.
    for row in rows:
        item_name = normalize_text(
            read_cell(
                ws_values,
                ws_write,
                row,
                columns["Tên hàng"],
            )
        )

        if item_name == normalized_discount_name:
            discount_rows.append(row)

    if not discount_rows:
        return None

    discount_row_set = set(discount_rows)

    # Nhóm dòng chiết khấu theo thuế suất.
    discount_rows_by_tax: dict[
        tuple[str, str],
        list[int],
    ] = defaultdict(list)

    for row in discount_rows:
        rate_value = read_cell(
            ws_values,
            ws_write,
            row,
            columns["Thuế suất"],
        )

        discount_rows_by_tax[
            tax_rate_key(rate_value)
        ].append(row)

    # Nhóm dòng sản phẩm theo thuế suất.
    product_rows_by_tax: dict[
        tuple[str, str],
        list[tuple[int, Decimal]],
    ] = defaultdict(list)

    for row in rows:
        if row in discount_row_set:
            continue

        item_name = normalize_text(
            read_cell(
                ws_values,
                ws_write,
                row,
                columns["Tên hàng"],
            )
        )

        if not item_name:
            continue

        quantity = to_decimal(
            read_cell(
                ws_values,
                ws_write,
                row,
                columns["Số lượng"],
            )
        ) or Decimal("0")

        unit_price = to_decimal(
            read_cell(
                ws_values,
                ws_write,
                row,
                columns["Đơn giá"],
            )
        ) or Decimal("0")

        # Doanh thu trước chiết khấu.
        gross_revenue = round_money(
            quantity * unit_price
        )

        if gross_revenue <= 0:
            continue

        rate_value = read_cell(
            ws_values,
            ws_write,
            row,
            columns["Thuế suất"],
        )

        rate_key = tax_rate_key(rate_value)

        product_rows_by_tax[rate_key].append(
            (
                row,
                gross_revenue,
            )
        )

    total_discount = Decimal("0")
    total_allocated = Decimal("0")
    total_remaining = Decimal("0")
    affected_products = 0

    # Xử lý riêng từng nhóm thuế suất.
    for rate_key, tax_discount_rows in (
        discount_rows_by_tax.items()
    ):
        discount_total = Decimal("0")

        # Tổng chiết khấu của đúng thuế suất này.
        for row in tax_discount_rows:
            discount_revenue = to_decimal(
                read_cell(
                    ws_values,
                    ws_write,
                    row,
                    columns["Doanh thu"],
                )
            ) or Decimal("0")

            discount_total += abs(
                discount_revenue
            )

        discount_total = round_money(
            discount_total
        )

        if discount_total <= 0:
            continue

        total_discount += discount_total

        # Chỉ lấy sản phẩm cùng thuế suất.
        product_rows = list(
            product_rows_by_tax.get(
                rate_key,
                [],
            )
        )

        # Ưu tiên Số lượng × Đơn giá lớn nhất.
        product_rows.sort(
            key=lambda item: (
                -item[1],
                item[0],
            )
        )

        remaining_discount = discount_total

        for row, gross_revenue in product_rows:
            if remaining_discount <= 0:
                break

            old_discount = to_decimal(
                read_cell(
                    ws_values,
                    ws_write,
                    row,
                    columns["Chiết khấu"],
                )
            ) or Decimal("0")

            old_tax = to_decimal(
                read_cell(
                    ws_values,
                    ws_write,
                    row,
                    columns["Tiền thuế"],
                )
            ) or Decimal("0")

            rate_value = read_cell(
                ws_values,
                ws_write,
                row,
                columns["Thuế suất"],
            )

            allocated_discount = min(
                gross_revenue,
                remaining_discount,
            )

            allocated_discount = round_money(
                allocated_discount
            )

            if allocated_discount <= 0:
                continue

            new_discount = round_money(
                old_discount
                + allocated_discount
            )

            new_revenue = round_money(
                gross_revenue
                - allocated_discount
            )

            rate = tax_ratio(
                rate_value=rate_value,
                old_revenue=gross_revenue,
                old_tax=old_tax,
            )

            new_tax = round_money(
                new_revenue * rate
            )

            new_total = round_money(
                new_revenue + new_tax
            )

            # Không thay đổi cột Đơn giá.
            set_money(
                ws_write,
                row,
                columns["Chiết khấu"],
                new_discount,
            )

            set_money(
                ws_write,
                row,
                columns["Doanh thu"],
                new_revenue,
            )

            set_money(
                ws_write,
                row,
                columns["Tiền thuế"],
                new_tax,
            )

            set_money(
                ws_write,
                row,
                columns["Thành tiền"],
                new_total,
            )

            remaining_discount = round_money(
                remaining_discount
                - allocated_discount
            )

            affected_products += 1

        allocated_for_tax = round_money(
            discount_total
            - remaining_discount
        )

        total_allocated += allocated_for_tax
        total_remaining += remaining_discount

        # Đưa dòng chiết khấu về 0 nếu đã phân bổ hết.
        # Nếu chưa phân bổ hết thì giữ phần còn lại ở dòng đầu.
        for index, row in enumerate(
            tax_discount_rows
        ):
            old_revenue = to_decimal(
                read_cell(
                    ws_values,
                    ws_write,
                    row,
                    columns["Doanh thu"],
                )
            ) or Decimal("0")

            old_tax = to_decimal(
                read_cell(
                    ws_values,
                    ws_write,
                    row,
                    columns["Tiền thuế"],
                )
            ) or Decimal("0")

            rate_value = read_cell(
                ws_values,
                ws_write,
                row,
                columns["Thuế suất"],
            )

            if (
                index == 0
                and remaining_discount > 0
            ):
                rate = tax_ratio(
                    rate_value=rate_value,
                    old_revenue=old_revenue,
                    old_tax=old_tax,
                )

                remaining_revenue = (
                    -remaining_discount
                )

                remaining_tax = round_money(
                    remaining_revenue * rate
                )

                remaining_total = round_money(
                    remaining_revenue
                    + remaining_tax
                )

                # Không sửa Đơn giá.
                set_money(
                    ws_write,
                    row,
                    columns["Chiết khấu"],
                    Decimal("0"),
                )

                set_money(
                    ws_write,
                    row,
                    columns["Doanh thu"],
                    remaining_revenue,
                )

                set_money(
                    ws_write,
                    row,
                    columns["Tiền thuế"],
                    remaining_tax,
                )

                set_money(
                    ws_write,
                    row,
                    columns["Thành tiền"],
                    remaining_total,
                )

            else:
                # Dòng chiết khấu đã được phân bổ xuống hàng.
                # Không sửa Đơn giá.
                set_money(
                    ws_write,
                    row,
                    columns["Chiết khấu"],
                    Decimal("0"),
                )

                set_money(
                    ws_write,
                    row,
                    columns["Doanh thu"],
                    Decimal("0"),
                )

                set_money(
                    ws_write,
                    row,
                    columns["Tiền thuế"],
                    Decimal("0"),
                )

                set_money(
                    ws_write,
                    row,
                    columns["Thành tiền"],
                    Decimal("0"),
                )

    if total_discount <= 0:
        return None

    return InvoiceResult(
        invoice=invoice,
        discount=round_money(
            total_discount
        ),
        allocated=round_money(
            total_allocated
        ),
        remaining=round_money(
            total_remaining
        ),
        affected_products=affected_products,
    )


def process_sheet(
    ws_values: Worksheet,
    ws_write: Worksheet,
) -> list[InvoiceResult]:
    """Xử lý toàn bộ hóa đơn trong một sheet."""

    header_info = find_header_row(ws_write)

    if header_info is None:
        return []

    header_row, columns = header_info

    invoice_rows: dict[
        tuple[str, ...],
        list[int],
    ] = defaultdict(list)

    for row in range(
        header_row + 1,
        ws_write.max_row + 1,
    ):
        key = build_invoice_key(
            ws_values=ws_values,
            ws_write=ws_write,
            row=row,
            columns=columns,
        )

        if key is not None:
            invoice_rows[key].append(row)

    results: list[InvoiceResult] = []

    for key, rows in invoice_rows.items():
        result = process_invoice(
            ws_values=ws_values,
            ws_write=ws_write,
            rows=rows,
            columns=columns,
            invoice=invoice_label(key),
        )

        if result is not None:
            results.append(result)

    return results


def main() -> None:
    parser = argparse.ArgumentParser(
        description=(
            "Phân bổ dòng 'Chiết khấu hàng hóa' vào "
            "các mặt hàng có đơn giá cao nhất trước."
        )
    )

    parser.add_argument(
        "input_file",
        help="Đường dẫn file Excel đầu vào (.xlsx hoặc .xlsm)",
    )

    parser.add_argument(
        "output_file",
        help="Đường dẫn file Excel đầu ra",
    )

    parser.add_argument(
        "--sheet",
        help=(
            "Chỉ xử lý một sheet cụ thể. "
            "Mặc định xử lý tất cả sheet có đúng tiêu đề."
        ),
    )

    args = parser.parse_args()

    input_path = Path(
        args.input_file
    ).resolve()

    output_path = Path(
        args.output_file
    ).resolve()

    if input_path == output_path:
        raise ValueError(
            "File đầu ra phải khác file đầu vào "
            "để tránh mất dữ liệu gốc."
        )

    if not input_path.exists():
        raise FileNotFoundError(
            f"Không tìm thấy file: {input_path}"
        )

    if input_path.suffix.lower() not in {
        ".xlsx",
        ".xlsm",
    }:
        raise ValueError(
            "Chương trình chỉ hỗ trợ file .xlsx và .xlsm."
        )

    output_path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    keep_vba = (
        input_path.suffix.lower() == ".xlsm"
    )

    # Workbook dùng để ghi và giữ nguyên công thức/định dạng.
    wb_write = load_workbook(
        input_path,
        data_only=False,
        keep_vba=keep_vba,
    )

    # Workbook dùng để đọc kết quả đã tính của công thức.
    wb_values = load_workbook(
        input_path,
        data_only=True,
        keep_vba=keep_vba,
    )

    if args.sheet:
        if args.sheet not in wb_write.sheetnames:
            raise ValueError(
                f"Không tìm thấy sheet: {args.sheet}"
            )

        sheet_names = [args.sheet]

    else:
        sheet_names = wb_write.sheetnames

    all_results: list[
        tuple[str, InvoiceResult]
    ] = []

    sheets_found = 0

    for sheet_name in sheet_names:
        ws_write = wb_write[sheet_name]
        ws_values = wb_values[sheet_name]

        if find_header_row(ws_write) is None:
            print(
                f"Bỏ qua sheet '{sheet_name}': "
                "không tìm thấy đầy đủ các cột yêu cầu."
            )
            continue

        sheets_found += 1

        results = process_sheet(
            ws_values=ws_values,
            ws_write=ws_write,
        )

        all_results.extend(
            (sheet_name, result)
            for result in results
        )

    if sheets_found == 0:
        raise ValueError(
            "Không có sheet nào chứa đầy đủ các cột: "
            + ", ".join(REQUIRED_HEADERS)
        )

    wb_write.save(output_path)

    total_discount = sum(
        (
            result.discount
            for _, result in all_results
        ),
        Decimal("0"),
    )

    total_allocated = sum(
        (
            result.allocated
            for _, result in all_results
        ),
        Decimal("0"),
    )

    total_remaining = sum(
        (
            result.remaining
            for _, result in all_results
        ),
        Decimal("0"),
    )

    print()
    print(f"Đã xuất file: {output_path}")
    print(
        "Số hóa đơn có dòng chiết khấu đã xử lý: "
        f"{len(all_results)}"
    )
    print(
        f"Tổng chiết khấu: "
        f"{excel_number(total_discount):,}"
    )
    print(
        f"Đã phân bổ: "
        f"{excel_number(total_allocated):,}"
    )
    print(
        f"Chưa phân bổ: "
        f"{excel_number(total_remaining):,}"
    )

    print()

    for sheet_name, result in all_results:
        status = (
            "OK"
            if result.remaining == 0
            else "CÒN DƯ"
        )

        print(
            f"[{status}] "
            f"{sheet_name} | "
            f"{result.invoice} | "
            f"CK={excel_number(result.discount):,} | "
            f"Phân bổ={excel_number(result.allocated):,} | "
            f"Còn={excel_number(result.remaining):,} | "
            f"Số mặt hàng bị trừ="
            f"{result.affected_products}"
        )


if __name__ == "__main__":
    main()